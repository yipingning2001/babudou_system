import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, date, timedelta
import openpyxl
from database import get_db
import models
from auth import require_owner

router = APIRouter(prefix="/members", tags=["会员"])

EXCEL_HEADERS = ["手机号", "家长姓名", "孩子姓名", "孩子生日(YYYY-MM-DD)", "当前鞋码"]


class MemberCreate(BaseModel):
    phone: str
    name: Optional[str] = None
    child_name: Optional[str] = None
    child_birthday: Optional[datetime] = None
    current_size: Optional[str] = None
    register_store_id: Optional[int] = None


class MemberOut(BaseModel):
    id: int
    phone: str
    name: Optional[str]
    child_name: Optional[str]
    child_birthday: Optional[datetime]
    current_size: Optional[str]
    points: int
    total_spent: float

    class Config:
        from_attributes = True


@router.get("/", response_model=list[MemberOut])
def list_members(store_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(models.Member)
    if store_id:
        query = query.filter(models.Member.register_store_id == store_id)
    return query.order_by(models.Member.total_spent.desc()).all()


@router.get("/search")
def search_member(phone: str, db: Session = Depends(get_db)):
    """按手机号查找会员（收银时用）"""
    member = db.query(models.Member).filter(models.Member.phone == phone).first()
    if not member:
        raise HTTPException(status_code=404, detail="会员不存在")
    return member


@router.post("/", response_model=MemberOut)
def create_member(member: MemberCreate, db: Session = Depends(get_db)):
    existing = db.query(models.Member).filter(models.Member.phone == member.phone).first()
    if existing:
        raise HTTPException(status_code=400, detail="该手机号已注册")
    db_member = models.Member(**member.model_dump())
    db.add(db_member)
    db.commit()
    db.refresh(db_member)
    return db_member


@router.put("/{member_id}", response_model=MemberOut)
def update_member(member_id: int, member: MemberCreate, db: Session = Depends(get_db)):
    db_member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="会员不存在")
    for key, value in member.model_dump(exclude_unset=True).items():
        setattr(db_member, key, value)
    db.commit()
    db.refresh(db_member)
    return db_member


@router.get("/birthday-reminder")
def birthday_reminder(days_ahead: int = 7, db: Session = Depends(get_db)):
    """查询未来N天内孩子生日的会员，用于营销回访"""
    today = date.today()
    results = []
    members = db.query(models.Member).filter(models.Member.child_birthday.isnot(None)).all()
    for m in members:
        bday = m.child_birthday
        this_year_bday = bday.replace(year=today.year)
        if this_year_bday.date() < today:
            this_year_bday = bday.replace(year=today.year + 1)
        delta = (this_year_bday.date() - today).days
        if 0 <= delta <= days_ahead:
            results.append({
                "member_id": m.id,
                "name": m.name,
                "phone": m.phone,
                "child_name": m.child_name,
                "birthday": this_year_bday.strftime("%m-%d"),
                "days_until": delta,
                "current_size": m.current_size,
            })
    return sorted(results, key=lambda x: x["days_until"])


@router.get("/import-template")
def download_import_template(_owner: models.User = Depends(require_owner)):
    """下载会员导入用的Excel模板（带表头+示例行）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "会员导入模板"
    ws.append(EXCEL_HEADERS)
    ws.append(["13800000001", "王女士", "小明", "2020-05-01", "26"])

    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=member_import_template.xlsx"},
    )


@router.post("/import-excel")
def import_members_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _owner: models.User = Depends(require_owner),
):
    """
    按模板格式批量导入会员。
    列顺序：手机号(必填) / 家长姓名 / 孩子姓名 / 孩子生日(YYYY-MM-DD) / 当前鞋码
    手机号已存在的行会被跳过（不覆盖已有数据），并在结果里列出来。
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 .xlsx 或 .xls 格式的文件")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="文件解析失败，请确认是有效的Excel文件")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头

    created, skipped_existing, skipped_invalid = [], [], []

    for row_idx, row in enumerate(rows, start=2):
        if not row or all(cell is None for cell in row):
            continue
        phone = str(row[0]).strip() if row[0] is not None else ""
        if not phone:
            skipped_invalid.append({"row": row_idx, "reason": "手机号为空"})
            continue

        if db.query(models.Member).filter(models.Member.phone == phone).first():
            skipped_existing.append({"row": row_idx, "phone": phone})
            continue

        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        child_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None

        child_birthday = None
        raw_birthday = row[3] if len(row) > 3 else None
        if raw_birthday:
            if isinstance(raw_birthday, datetime):
                child_birthday = raw_birthday
            else:
                try:
                    child_birthday = datetime.strptime(str(raw_birthday).strip(), "%Y-%m-%d")
                except ValueError:
                    skipped_invalid.append({"row": row_idx, "reason": f"生日格式不对：{raw_birthday}，应为 YYYY-MM-DD"})
                    continue

        current_size = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None

        db.add(models.Member(
            phone=phone,
            name=name,
            child_name=child_name,
            child_birthday=child_birthday,
            current_size=current_size,
        ))
        created.append(phone)

    db.commit()

    return {
        "created_count": len(created),
        "skipped_existing_count": len(skipped_existing),
        "skipped_invalid_count": len(skipped_invalid),
        "skipped_existing": skipped_existing,
        "skipped_invalid": skipped_invalid,
        "message": f"导入完成：新增 {len(created)} 个，跳过已存在 {len(skipped_existing)} 个，跳过无效数据 {len(skipped_invalid)} 个",
    }


@router.get("/export-excel")
def export_members_excel(db: Session = Depends(get_db), _owner: models.User = Depends(require_owner)):
    """导出全部会员数据为Excel"""
    members = db.query(models.Member).order_by(models.Member.total_spent.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "会员列表"
    headers = ["手机号", "家长姓名", "孩子姓名", "孩子生日", "当前鞋码", "积分", "累计消费", "注册门店"]
    ws.append(headers)

    for m in members:
        ws.append([
            m.phone,
            m.name,
            m.child_name,
            m.child_birthday.strftime("%Y-%m-%d") if m.child_birthday else None,
            m.current_size,
            m.points,
            m.total_spent,
            m.register_store.name if m.register_store else None,
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    today_str = date.today().isoformat()
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=members_export_{today_str}.xlsx"},
    )


@router.delete("/{member_id}")
def delete_member(member_id: int, db: Session = Depends(get_db)):
    db_member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="会员不存在")
    has_orders = db.query(models.Order).filter(models.Order.member_id == member_id).first()
    if has_orders:
        raise HTTPException(status_code=400, detail="该会员有历史订单，无法删除，以免影响账目记录")
    db.delete(db_member)
    db.commit()
    return {"message": "已删除"}


@router.get("/{member_id}/orders")
def member_orders(member_id: int, db: Session = Depends(get_db)):
    """会员历史购买记录"""
    member = db.query(models.Member).filter(models.Member.id == member_id).first()
    if not member:
        raise HTTPException(status_code=404, detail="会员不存在")
    orders = (
        db.query(models.Order)
        .filter(models.Order.member_id == member_id)
        .order_by(models.Order.created_at.desc())
        .all()
    )
    return {
        "member": {"name": member.name, "phone": member.phone, "points": member.points, "total_spent": member.total_spent},
        "orders": [
            {
                "order_id": o.id,
                "date": o.created_at.strftime("%Y-%m-%d %H:%M"),
                "amount": o.total_amount,
                "items": [
                    {
                        "product": f"{item.product.model_no} {item.product.name}",
                        "color": item.product.color,
                        "size": item.size,
                        "qty": item.quantity,
                        "price": item.unit_price,
                    }
                    for item in o.items
                ],
            }
            for o in orders
        ],
    }
